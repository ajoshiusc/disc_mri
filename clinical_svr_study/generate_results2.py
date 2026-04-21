#!/usr/bin/env python3
"""Reproduce manuscript tables and figure source files from SVR2.xlsx.

This script:
1. Reads the paired raw-stack and SVR review data from SVR2.xlsx.
2. Recovers the blinded SVR case mapping from the study repository permutation.
3. Computes descriptive and paired inferential summaries.
4. Writes CSV files used by the LaTeX draft to render tables/figures.
5. Writes ready-to-input LaTeX tables for convenience.

The Wilcoxon signed-rank p-values are computed with a normal approximation and
tie correction so the project is self-contained and does not depend on SciPy.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import load_workbook


# Randomization permutation recovered from:
# ajoshiusc/disc_mri/clinical_svr_study/svr_8_3_2023/main_shuffle_rename.py
SVR_RANDOMIZATION = [
    26, 7, 13, 2, 30, 20, 18, 31, 12, 17, 1, 14, 9, 21, 11, 25,
    3, 28, 19, 15, 8, 5, 24, 6, 29, 10, 4, 23, 16, 22, 27,
]

INVERSE_RANDOMIZATION = {rnd: index + 1 for index, rnd in enumerate(SVR_RANDOMIZATION)}

METRIC_LABELS = {
    "conf": "Diagnostic confidence",
    "time": "Reading time (min)",
    "cc": "Corpus callosum visibility",
    "csp": "CSP visibility",
    "vent": "Ventricles / CSF visibility",
    "pf": "Posterior fossa visibility",
}

STRUCTURE_KEYS = ["cc", "csp", "vent", "pf"]
ALL_KEYS = ["conf", "time", "cc", "csp", "vent", "pf"]

PRIMARY_CATEGORY_ORDER = [
    "Normal intracranial MRI",
    "Ventriculomegaly",
    "Hemorrhage / porencephalic injury",
    "Callosal anomaly",
    "Chiari II / hindbrain anomaly",
    "Posterior fossa / cerebellar anomaly",
    "CSP anomaly",
    "Arachnoid cyst",
    "Holoprosencephaly",
    "Extra-cranial finding with normal brain",
    "Other",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to SVR2.xlsx",
    )
    parser.add_argument(
        "--outdir",
        default="outputs",
        help="Output directory for CSV and LaTeX artifacts",
    )
    return parser.parse_args()


def fmt_float(value: float, digits: int = 2) -> str:
    return f"{value:.{digits}f}"


def fmt_pvalue(value: float) -> str:
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def normal_cdf(z: float) -> float:
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def wilcoxon_signed_rank(raw: Sequence[float], svr: Sequence[float]) -> Dict[str, float]:
    """Approximate two-sided Wilcoxon signed-rank test with tie correction."""
    diffs = [sv - rv for rv, sv in zip(raw, svr)]
    nonzero = [diff for diff in diffs if diff != 0]
    if not nonzero:
        return {
            "n_nonzero": 0,
            "w_plus": 0.0,
            "w_minus": 0.0,
            "z_approx": 0.0,
            "p_approx": 1.0,
        }

    abs_values = [abs(diff) for diff in nonzero]
    sorted_indices = sorted(range(len(abs_values)), key=lambda idx: abs_values[idx])
    ranks = [0.0] * len(abs_values)
    tie_counts: List[int] = []
    start = 0
    while start < len(abs_values):
        end = start
        while end + 1 < len(abs_values) and abs_values[sorted_indices[end + 1]] == abs_values[sorted_indices[start]]:
            end += 1
        avg_rank = ((start + 1) + (end + 1)) / 2.0
        for pos in range(start, end + 1):
            ranks[sorted_indices[pos]] = avg_rank
        tie_counts.append(end - start + 1)
        start = end + 1

    w_plus = sum(rank for diff, rank in zip(nonzero, ranks) if diff > 0)
    w_minus = sum(rank for diff, rank in zip(nonzero, ranks) if diff < 0)
    n = len(nonzero)
    mean_w = n * (n + 1) / 4.0
    tie_term = sum(t * (t + 1) * (2 * t + 1) for t in tie_counts if t > 1)
    var_w = (n * (n + 1) * (2 * n + 1) - tie_term / 2.0) / 24.0

    if var_w <= 0:
        z_value = 0.0
        p_value = 1.0
    else:
        continuity = 0.5 if w_plus > mean_w else -0.5 if w_plus < mean_w else 0.0
        z_value = (w_plus - mean_w - continuity) / math.sqrt(var_w)
        p_value = 2.0 * (1.0 - normal_cdf(abs(z_value)))

    return {
        "n_nonzero": n,
        "w_plus": w_plus,
        "w_minus": w_minus,
        "z_approx": z_value,
        "p_approx": p_value,
    }


def is_nondiagnostic(text: str) -> bool:
    lower = text.lower()
    return any(token in lower for token in ("xxxx", "poorly visualized", "incomplete"))


def diagnosis_sets(text: str) -> List[str]:
    lower = text.lower()
    categories: List[str] = []

    abnormal_terms = [
        "ventric",
        "agenesis",
        "colpocephaly",
        "holoprosenceph",
        "hindbrain",
        "chiari",
        "cerebell",
        "cyst",
        "hemorrhage",
        "haemorrhage",
        "csp",
        "malformation",
        "aqueductal",
    ]

    if "nl" in lower and not any(term in lower for term in abnormal_terms):
        categories.append("Normal intracranial MRI")
    if "ventric" in lower or "aqueductal" in lower:
        categories.append("Ventriculomegaly")
    if "cc agenesis" in lower or "colpocephaly" in lower:
        categories.append("Callosal anomaly")
    if "holoprosenceph" in lower:
        categories.append("Holoprosencephaly")
    if "hindbrain" in lower or "chiari" in lower:
        categories.append("Chiari II / hindbrain anomaly")
    if (
        "small cerebellum" in lower
        or "pf cyst" in lower
        or re.search(r"\bpf\b", lower)
        or "cerebell" in lower
        or "mega cisterna" in lower
    ):
        categories.append("Posterior fossa / cerebellar anomaly")
    if "porenceph" in lower or "hemorrhage" in lower or "haemorrhage" in lower:
        categories.append("Hemorrhage / porencephalic injury")
    if "arachnoid cyst" in lower:
        categories.append("Arachnoid cyst")
    if "csp" in lower and "perforated csp" not in lower:
        categories.append("CSP anomaly")
    if "ear malformation" in lower:
        categories.append("Extra-cranial finding with normal brain")
    if not categories:
        categories.append("Other")

    deduped: List[str] = []
    for category in categories:
        if category not in deduped:
            deduped.append(category)
    return deduped


def primary_category(text: str) -> str:
    category_set = diagnosis_sets(text)
    for category in PRIMARY_CATEGORY_ORDER:
        if category in category_set:
            return category
    return category_set[0]


def normal_intracranial(text: str) -> bool:
    return "Normal intracranial MRI" in diagnosis_sets(text)


def broad_agreement(raw_text: str, svr_text: str) -> bool:
    family_map = {
        "Chiari II / hindbrain anomaly": "Posterior fossa / hindbrain family",
        "Posterior fossa / cerebellar anomaly": "Posterior fossa / hindbrain family",
    }
    raw_set = {family_map.get(item, item) for item in diagnosis_sets(raw_text)}
    svr_set = {family_map.get(item, item) for item in diagnosis_sets(svr_text)}
    return bool(raw_set & svr_set)


def load_cases(xlsx_path: Path) -> List[Dict[str, object]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook["Sheet1"]

    raw_rows: Dict[int, Dict[str, object]] = {}
    for row in range(3, 34):
        case_id = sheet.cell(row, 2).value
        if case_id is None:
            continue
        raw_rows[int(case_id)] = {
            "case_id": int(case_id),
            "raw_dx": str(sheet.cell(row, 3).value),
            "raw_conf": int(sheet.cell(row, 4).value),
            "raw_time": float(sheet.cell(row, 5).value),
            "raw_cc": int(sheet.cell(row, 6).value),
            "raw_csp": int(sheet.cell(row, 7).value),
            "raw_vent": int(sheet.cell(row, 8).value),
            "raw_pf": int(sheet.cell(row, 9).value),
        }

    svr_rows: Dict[int, Dict[str, object]] = {}
    for row in range(36, 67):
        randomized_id = sheet.cell(row, 2).value
        if randomized_id is None:
            continue
        original_case_id = INVERSE_RANDOMIZATION[int(randomized_id)]
        diagnosis = str(sheet.cell(row, 3).value)
        svr_rows[original_case_id] = {
            "svr_randomized_id": int(randomized_id),
            "svr_dx": diagnosis,
            "svr_conf": int(sheet.cell(row, 4).value),
            "svr_time": float(sheet.cell(row, 5).value),
            "svr_cc": int(sheet.cell(row, 6).value),
            "svr_csp": int(sheet.cell(row, 7).value),
            "svr_vent": int(sheet.cell(row, 8).value),
            "svr_pf": int(sheet.cell(row, 9).value),
            "svr_nondiagnostic": is_nondiagnostic(diagnosis),
        }

    paired_cases: List[Dict[str, object]] = []
    for case_id in sorted(raw_rows):
        merged = dict(raw_rows[case_id])
        merged.update(svr_rows[case_id])
        merged["raw_category"] = primary_category(str(merged["raw_dx"]))
        merged["svr_primary_category"] = primary_category(str(merged["svr_dx"]))
        merged["raw_categories"] = "; ".join(diagnosis_sets(str(merged["raw_dx"])))
        merged["svr_categories"] = "; ".join(diagnosis_sets(str(merged["svr_dx"])))
        merged["normal_abnormal_agreement"] = (
            normal_intracranial(str(merged["raw_dx"])) == normal_intracranial(str(merged["svr_dx"]))
        )
        merged["broad_category_agreement"] = broad_agreement(str(merged["raw_dx"]), str(merged["svr_dx"]))
        paired_cases.append(merged)
    return paired_cases


def mean(values: Sequence[float]) -> float:
    return float(statistics.mean(values))


def sd(values: Sequence[float]) -> float:
    if len(values) <= 1:
        return 0.0
    return float(statistics.stdev(values))


def median(values: Sequence[float]) -> float:
    return float(statistics.median(values))


def summarize_subset(cases: Sequence[Dict[str, object]], subset_name: str) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for key in ALL_KEYS:
        raw_values = [float(case[f"raw_{key}"]) for case in cases]
        svr_values = [float(case[f"svr_{key}"]) for case in cases]
        deltas = [sv - rv for rv, sv in zip(raw_values, svr_values)]
        test = wilcoxon_signed_rank(raw_values, svr_values)
        rows.append(
            {
                "subset": subset_name,
                "metric_key": key,
                "metric_label": METRIC_LABELS[key],
                "n": len(cases),
                "raw_mean": mean(raw_values),
                "raw_sd": sd(raw_values),
                "raw_median": median(raw_values),
                "svr_mean": mean(svr_values),
                "svr_sd": sd(svr_values),
                "svr_median": median(svr_values),
                "mean_delta_svr_minus_raw": mean(deltas),
                "median_delta_svr_minus_raw": median(deltas),
                **test,
            }
        )
    return rows


def write_csv(path: Path, rows: Iterable[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def latex_escape(text: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def render_metrics_table(summary_rows: Sequence[Dict[str, object]], caption: str, label: str) -> str:
    lines = [
        r"\begin{table}[htbp]",
        r"\centering",
        rf"\caption{{{caption}}}",
        rf"\label{{{label}}}",
        r"\begin{threeparttable}",
        r"\begin{tabular}{lcccccc}",
        r"\toprule",
        r"Metric & Raw mean $\pm$ SD & Raw median & SVR mean $\pm$ SD & SVR median & Mean $\Delta$ & $p$ \\",
        r"\midrule",
    ]
    for row in summary_rows:
        lines.append(
            r"{} & {} $\pm$ {} & {} & {} $\pm$ {} & {} & {} & {} \\".format(
                latex_escape(str(row["metric_label"])),
                fmt_float(float(row["raw_mean"])),
                fmt_float(float(row["raw_sd"])),
                fmt_float(float(row["raw_median"]), 0),
                fmt_float(float(row["svr_mean"])),
                fmt_float(float(row["svr_sd"])),
                fmt_float(float(row["svr_median"]), 0),
                fmt_float(float(row["mean_delta_svr_minus_raw"])),
                fmt_pvalue(float(row["p_approx"])),
            )
        )
    lines.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            r"\begin{tablenotes}[flushleft]",
            r"\footnotesize",
            r"\item Mean $\Delta$ denotes SVR minus raw-stack values. P-values are approximate Wilcoxon signed-rank tests computed by the accompanying reproducibility script.",
            r"\end{tablenotes}",
            r"\end{threeparttable}",
            r"\end{table}",
        ]
    )
    return "\n".join(lines) + "\n"


def render_distribution_table(distribution_rows: Sequence[Dict[str, object]]) -> str:
    lines = [
        r"\begin{table}[htbp]",
        r"\centering",
        r"\caption{Case mix based on the raw-stack diagnosis labels.}",
        r"\label{tab:cohort_distribution}",
        r"\begin{tabular}{lc}",
        r"\toprule",
        r"Diagnostic category & Cases \\",
        r"\midrule",
    ]
    for row in distribution_rows:
        lines.append(f"{latex_escape(str(row['category']))} & {row['count']} \\\\")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}"])
    return "\n".join(lines) + "\n"


def render_diagnostic_summary_table(summary_rows: Sequence[Dict[str, object]]) -> str:
    lines = [
        r"\begin{table}[htbp]",
        r"\centering",
        r"\caption{Heuristic diagnostic concordance summary derived from the paired free-text reads.}",
        r"\label{tab:diagnostic_summary}",
        r"\begin{tabular}{lcc}",
        r"\toprule",
        r"Summary measure & Count & Proportion (\%) \\",
        r"\midrule",
    ]
    for row in summary_rows:
        lines.append(
            "{} & {} & {} \\\\".format(
                latex_escape(str(row["measure"])),
                row["count"],
                fmt_float(float(row["proportion"]) * 100.0, 1),
            )
        )
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}"])
    return "\n".join(lines) + "\n"


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    outdir = Path(args.outdir).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    cases = load_cases(xlsx_path)
    all_summary = summarize_subset(cases, "all_cases")
    clean_cases = [case for case in cases if not bool(case["svr_nondiagnostic"])]
    clean_summary = summarize_subset(clean_cases, "excluding_nondiagnostic_svr")

    paired_case_fieldnames = [
        "case_id",
        "svr_randomized_id",
        "raw_dx",
        "svr_dx",
        "raw_category",
        "svr_primary_category",
        "raw_categories",
        "svr_categories",
        "normal_abnormal_agreement",
        "broad_category_agreement",
        "svr_nondiagnostic",
        "raw_conf",
        "svr_conf",
        "raw_time",
        "svr_time",
        "raw_cc",
        "svr_cc",
        "raw_csp",
        "svr_csp",
        "raw_vent",
        "svr_vent",
        "raw_pf",
        "svr_pf",
    ]
    write_csv(outdir / "paired_case_table.csv", cases, paired_case_fieldnames)

    summary_fieldnames = [
        "subset",
        "metric_key",
        "metric_label",
        "n",
        "raw_mean",
        "raw_sd",
        "raw_median",
        "svr_mean",
        "svr_sd",
        "svr_median",
        "mean_delta_svr_minus_raw",
        "median_delta_svr_minus_raw",
        "n_nonzero",
        "w_plus",
        "w_minus",
        "z_approx",
        "p_approx",
    ]
    write_csv(outdir / "summary_stats.csv", all_summary + clean_summary, summary_fieldnames)

    outcome_plot_rows = []
    for row in all_summary:
        if row["metric_key"] in ("conf", "time"):
            outcome_plot_rows.append(
                {
                    "metric": row["metric_label"],
                    "raw_mean": row["raw_mean"],
                    "raw_sd": row["raw_sd"],
                    "svr_mean": row["svr_mean"],
                    "svr_sd": row["svr_sd"],
                }
            )
    write_csv(
        outdir / "outcome_bar_chart.csv",
        outcome_plot_rows,
        ["metric", "raw_mean", "raw_sd", "svr_mean", "svr_sd"],
    )

    structure_plot_rows = []
    for row in all_summary:
        if row["metric_key"] in STRUCTURE_KEYS:
            structure_plot_rows.append(
                {
                    "structure": row["metric_label"].replace(" visibility", ""),
                    "raw_mean": row["raw_mean"],
                    "raw_sd": row["raw_sd"],
                    "svr_mean": row["svr_mean"],
                    "svr_sd": row["svr_sd"],
                }
            )
    write_csv(
        outdir / "structure_bar_chart.csv",
        structure_plot_rows,
        ["structure", "raw_mean", "raw_sd", "svr_mean", "svr_sd"],
    )

    cohort_counts = Counter(str(case["raw_category"]) for case in cases)
    distribution_rows = [
        {"category": category, "count": cohort_counts.get(category, 0)}
        for category in PRIMARY_CATEGORY_ORDER
        if cohort_counts.get(category, 0) > 0
    ]
    write_csv(outdir / "cohort_distribution.csv", distribution_rows, ["category", "count"])

    total_cases = len(cases)
    diagnostic_summary_rows = [
        {
            "measure": "Normal/abnormal agreement",
            "count": sum(bool(case["normal_abnormal_agreement"]) for case in cases),
            "proportion": sum(bool(case["normal_abnormal_agreement"]) for case in cases) / total_cases,
        },
        {
            "measure": "Broad category agreement",
            "count": sum(bool(case["broad_category_agreement"]) for case in cases),
            "proportion": sum(bool(case["broad_category_agreement"]) for case in cases) / total_cases,
        },
        {
            "measure": "SVR limited/non-diagnostic cases",
            "count": sum(bool(case["svr_nondiagnostic"]) for case in cases),
            "proportion": sum(bool(case["svr_nondiagnostic"]) for case in cases) / total_cases,
        },
        {
            "measure": "Cases retained after excluding limited SVR reconstructions",
            "count": len(clean_cases),
            "proportion": len(clean_cases) / total_cases,
        },
    ]
    write_csv(
        outdir / "diagnostic_summary.csv",
        diagnostic_summary_rows,
        ["measure", "count", "proportion"],
    )

    (outdir / "table_metrics_all.tex").write_text(
        render_metrics_table(
            all_summary,
            "Paired comparison of reader-reported outcomes across all 31 cases.",
            "tab:paired_metrics_all",
        ),
        encoding="utf-8",
    )
    (outdir / "table_metrics_excluding_nondx.tex").write_text(
        render_metrics_table(
            clean_summary,
            "Sensitivity analysis excluding the four SVR reads marked as limited or non-diagnostic.",
            "tab:paired_metrics_clean",
        ),
        encoding="utf-8",
    )
    (outdir / "table_cohort_distribution.tex").write_text(
        render_distribution_table(distribution_rows),
        encoding="utf-8",
    )
    (outdir / "table_diagnostic_summary.tex").write_text(
        render_diagnostic_summary_table(diagnostic_summary_rows),
        encoding="utf-8",
    )

    snapshot = {
        "input_workbook": str(xlsx_path),
        "total_cases": total_cases,
        "nondiagnostic_svr_cases": [
            int(case["case_id"]) for case in cases if bool(case["svr_nondiagnostic"])
        ],
        "summary_stats": all_summary + clean_summary,
        "diagnostic_summary": diagnostic_summary_rows,
    }
    (outdir / "results_snapshot.json").write_text(
        json.dumps(snapshot, indent=2),
        encoding="utf-8",
    )

    print(f"Wrote reproducibility artifacts to {outdir}")


if __name__ == "__main__":
    main()
